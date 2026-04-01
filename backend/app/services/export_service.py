"""Export service for generating reports"""

import logging
import io
import csv
import json
from datetime import datetime
from typing import Optional, List, Dict, Any
from pathlib import Path

from sqlalchemy.orm import Session

from ..models.router import Router
from ..models.group import RouterGroup
from ..models.task import Task
from ..models.backup import RouterBackup
from ..models.monitoring import HealthCheck, AlertHistory
from ..schemas.report import ReportRequest, ReportFilter
from ..config import settings

logger = logging.getLogger(__name__)


class ExportService:
    """Service for generating and exporting reports"""

    def __init__(self, db: Session):
        self.db = db
        self.export_dir = settings.DATA_DIR / "exports"
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def generate_report(
        self,
        report_request: ReportRequest,
        user_id: Optional[int] = None
    ) -> Dict[str, Any]:
        """Generate a report based on request"""
        report_type = report_request.report_type
        format_type = report_request.format

        # Get data based on report type
        if report_type == "inventory":
            data = self._get_inventory_data(report_request.filters)
        elif report_type == "updates":
            data = self._get_update_data(report_request.filters)
        elif report_type == "health":
            data = self._get_health_data(report_request.filters)
        elif report_type == "activity":
            data = self._get_activity_data(report_request.filters)
        elif report_type == "backups":
            data = self._get_backup_data(report_request.filters)
        else:
            data = self._get_custom_data(report_request.filters, report_request.sections)

        # Generate output based on format
        if format_type == "pdf":
            return self._generate_pdf(report_request, data)
        elif format_type == "excel":
            return self._generate_excel(report_request, data)
        elif format_type == "csv":
            return self._generate_csv(report_request, data)
        else:  # json
            return self._generate_json(report_request, data)

    def _apply_filters(self, query, filters: ReportFilter, model=Router):
        """Apply common filters to query"""
        if filters.router_ids:
            query = query.filter(model.id.in_(filters.router_ids))

        if filters.status:
            if "online" in filters.status:
                query = query.filter(model.is_online == True)
            if "offline" in filters.status:
                query = query.filter(model.is_online == False)
            if "needs_update" in filters.status:
                query = query.filter(model.has_updates == True)

        if filters.date_from:
            query = query.filter(model.updated_at >= filters.date_from)
        if filters.date_to:
            query = query.filter(model.updated_at <= filters.date_to)

        return query

    def _get_inventory_data(self, filters: ReportFilter) -> Dict[str, Any]:
        """Get inventory report data"""
        query = self.db.query(Router)
        query = self._apply_filters(query, filters)
        routers = query.all()

        router_data = []
        version_summary = {}
        model_summary = {}

        for router in routers:
            groups = [g.name for g in router.groups] if router.groups else []

            router_data.append({
                "id": router.id,
                "ip": router.ip,
                "identity": router.identity,
                "model": router.model,
                "ros_version": router.ros_version,
                "firmware": router.firmware,
                "update_channel": router.update_channel,
                "is_online": router.is_online,
                "has_updates": router.has_updates,
                "location": router.location,
                "tags": router.tags or [],
                "groups": groups,
                "uptime": router.uptime,
                "last_seen": router.last_seen.isoformat() if router.last_seen else None
            })

            # Count versions
            version = router.ros_version or "Unknown"
            version_summary[version] = version_summary.get(version, 0) + 1

            # Count models
            model = router.model or "Unknown"
            model_summary[model] = model_summary.get(model, 0) + 1

        return {
            "generated_at": datetime.utcnow().isoformat(),
            "total_routers": len(routers),
            "online_count": sum(1 for r in routers if r.is_online),
            "offline_count": sum(1 for r in routers if not r.is_online),
            "needs_update_count": sum(1 for r in routers if r.has_updates),
            "routers": router_data,
            "version_summary": version_summary,
            "model_summary": model_summary
        }

    def _get_update_data(self, filters: ReportFilter) -> Dict[str, Any]:
        """Get update history data"""
        query = self.db.query(Task).filter(Task.type.in_(["update", "firmware_upgrade"]))

        if filters.date_from:
            query = query.filter(Task.created_at >= filters.date_from)
        if filters.date_to:
            query = query.filter(Task.created_at <= filters.date_to)

        tasks = query.order_by(Task.created_at.desc()).limit(1000).all()

        updates = []
        for task in tasks:
            result = task.results or {}
            updates.append({
                "router_id": task.config.get("router_id") if task.config else None,
                "router_identity": result.get("identity"),
                "router_ip": result.get("ip"),
                "update_type": task.type,
                "from_version": result.get("from_version"),
                "to_version": result.get("to_version"),
                "status": task.status,
                "started_at": task.created_at.isoformat(),
                "completed_at": task.completed_at.isoformat() if task.completed_at else None,
                "error_message": result.get("error")
            })

        return {
            "generated_at": datetime.utcnow().isoformat(),
            "date_range_start": filters.date_from.isoformat() if filters.date_from else None,
            "date_range_end": filters.date_to.isoformat() if filters.date_to else None,
            "total_updates": len(updates),
            "successful_updates": sum(1 for u in updates if u["status"] == "completed"),
            "failed_updates": sum(1 for u in updates if u["status"] == "failed"),
            "updates": updates
        }

    def _get_health_data(self, filters: ReportFilter) -> Dict[str, Any]:
        """Get health report data"""
        query = self.db.query(Router)
        query = self._apply_filters(query, filters)
        routers = query.all()

        health_data = []
        for router in routers:
            # Get health checks for this router
            checks = self.db.query(HealthCheck).filter(
                HealthCheck.router_id == router.id
            ).order_by(HealthCheck.checked_at.desc()).limit(100).all()

            # Get alerts
            alerts = self.db.query(AlertHistory).filter(
                AlertHistory.router_id == router.id,
                AlertHistory.status == "active"
            ).all()

            # Calculate uptime percentage
            total_checks = len(checks)
            online_checks = sum(1 for c in checks if c.is_online)
            uptime_pct = (online_checks / total_checks * 100) if total_checks > 0 else 0

            # Calculate average latency
            latencies = [c.latency_ms for c in checks if c.latency_ms is not None]
            avg_latency = sum(latencies) / len(latencies) if latencies else None

            # Determine status
            latest = checks[0] if checks else None
            status = latest.status if latest else "unknown"

            issues = []
            for alert in alerts:
                issues.append(alert.title)

            health_data.append({
                "router_id": router.id,
                "router_identity": router.identity,
                "router_ip": router.ip,
                "status": status,
                "uptime_percentage": round(uptime_pct, 2),
                "avg_latency_ms": round(avg_latency, 2) if avg_latency else None,
                "alerts_count": len(alerts),
                "issues": issues
            })

        healthy = sum(1 for h in health_data if h["status"] == "ok")
        warning = sum(1 for h in health_data if h["status"] == "warning")
        critical = sum(1 for h in health_data if h["status"] == "critical")

        return {
            "generated_at": datetime.utcnow().isoformat(),
            "date_range_start": filters.date_from.isoformat() if filters.date_from else None,
            "date_range_end": filters.date_to.isoformat() if filters.date_to else None,
            "total_routers": len(routers),
            "healthy_count": healthy,
            "warning_count": warning,
            "critical_count": critical,
            "avg_uptime_percentage": round(
                sum(h["uptime_percentage"] for h in health_data) / len(health_data), 2
            ) if health_data else 0,
            "routers": health_data
        }

    def _get_activity_data(self, filters: ReportFilter) -> Dict[str, Any]:
        """Get activity log data"""
        query = self.db.query(Task)

        if filters.date_from:
            query = query.filter(Task.created_at >= filters.date_from)
        if filters.date_to:
            query = query.filter(Task.created_at <= filters.date_to)

        tasks = query.order_by(Task.created_at.desc()).limit(500).all()

        activities = []
        for task in tasks:
            activities.append({
                "id": task.id,
                "type": task.type,
                "status": task.status,
                "progress": task.progress,
                "created_at": task.created_at.isoformat(),
                "updated_at": task.completed_at.isoformat() if task.completed_at else None,
                "config": task.config,
                "result": task.results
            })

        return {
            "generated_at": datetime.utcnow().isoformat(),
            "total_activities": len(activities),
            "activities": activities
        }

    def _get_backup_data(self, filters: ReportFilter) -> Dict[str, Any]:
        """Get backup report data"""
        query = self.db.query(RouterBackup)

        if filters.router_ids:
            query = query.filter(RouterBackup.router_id.in_(filters.router_ids))
        if filters.date_from:
            query = query.filter(RouterBackup.created_at >= filters.date_from)
        if filters.date_to:
            query = query.filter(RouterBackup.created_at <= filters.date_to)

        backups = query.order_by(RouterBackup.created_at.desc()).all()

        backup_data = []
        for backup in backups:
            backup_data.append({
                "id": backup.id,
                "router_id": backup.router_id,
                "name": backup.name,
                "backup_type": backup.backup_type,
                "file_size": backup.file_size,
                "status": backup.status,
                "router_version": backup.router_version,
                "router_model": backup.router_model,
                "router_identity": backup.router_identity,
                "created_at": backup.created_at.isoformat()
            })

        return {
            "generated_at": datetime.utcnow().isoformat(),
            "total_backups": len(backups),
            "total_size": sum(b.file_size or 0 for b in backups),
            "backups": backup_data
        }

    def _get_custom_data(self, filters: ReportFilter, sections: List[str] = None) -> Dict[str, Any]:
        """Get custom report data combining multiple sections"""
        data = {"generated_at": datetime.utcnow().isoformat()}

        sections = sections or ["inventory", "health"]

        if "inventory" in sections:
            data["inventory"] = self._get_inventory_data(filters)
        if "updates" in sections:
            data["updates"] = self._get_update_data(filters)
        if "health" in sections:
            data["health"] = self._get_health_data(filters)
        if "activity" in sections:
            data["activity"] = self._get_activity_data(filters)
        if "backups" in sections:
            data["backups"] = self._get_backup_data(filters)

        return data

    def _generate_pdf(self, request: ReportRequest, data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate PDF report"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet

            filename = f"report_{request.report_type}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
            filepath = self.export_dir / filename

            doc = SimpleDocTemplate(str(filepath), pagesize=A4)
            styles = getSampleStyleSheet()
            elements = []

            # Title
            title = request.title or f"{request.report_type.title()} Report"
            elements.append(Paragraph(title, styles['Title']))
            elements.append(Spacer(1, 12))
            elements.append(Paragraph(f"Generated: {data.get('generated_at', '')}", styles['Normal']))
            elements.append(Spacer(1, 24))

            # Add content based on report type
            if request.report_type == "inventory" and "routers" in data:
                self._add_inventory_to_pdf(elements, data, styles)
            elif request.report_type == "health" and "routers" in data:
                self._add_health_to_pdf(elements, data, styles)

            doc.build(elements)

            return {
                "id": filename,
                "report_type": request.report_type,
                "title": title,
                "format": "pdf",
                "status": "completed",
                "file_path": str(filepath),
                "file_size": filepath.stat().st_size,
                "download_url": f"/api/reports/download/{filename}",
                "created_at": datetime.utcnow().isoformat()
            }

        except ImportError:
            logger.warning("reportlab not installed, PDF generation disabled")
            return self._generate_json(request, data)

    def _add_inventory_to_pdf(self, elements, data, styles):
        """Add inventory data to PDF"""
        from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors

        # Summary
        elements.append(Paragraph("Summary", styles['Heading2']))
        elements.append(Paragraph(f"Total Routers: {data['total_routers']}", styles['Normal']))
        elements.append(Paragraph(f"Online: {data['online_count']}", styles['Normal']))
        elements.append(Paragraph(f"Offline: {data['offline_count']}", styles['Normal']))
        elements.append(Paragraph(f"Needs Update: {data['needs_update_count']}", styles['Normal']))
        elements.append(Spacer(1, 12))

        # Router table
        if data.get("routers"):
            elements.append(Paragraph("Router List", styles['Heading2']))
            table_data = [["IP", "Identity", "Model", "Version", "Status"]]
            for router in data["routers"][:50]:  # Limit to 50 rows
                table_data.append([
                    router["ip"],
                    router["identity"] or "-",
                    router["model"] or "-",
                    router["ros_version"] or "-",
                    "Online" if router["is_online"] else "Offline"
                ])

            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
            ]))
            elements.append(table)

    def _add_health_to_pdf(self, elements, data, styles):
        """Add health data to PDF"""
        from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors

        elements.append(Paragraph("Health Summary", styles['Heading2']))
        elements.append(Paragraph(f"Healthy: {data['healthy_count']}", styles['Normal']))
        elements.append(Paragraph(f"Warning: {data['warning_count']}", styles['Normal']))
        elements.append(Paragraph(f"Critical: {data['critical_count']}", styles['Normal']))
        elements.append(Spacer(1, 12))

    def _generate_excel(self, request: ReportRequest, data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate Excel report"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            filename = f"report_{request.report_type}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = self.export_dir / filename

            wb = Workbook()
            ws = wb.active
            ws.title = request.report_type.title()

            # Add data based on report type
            if "routers" in data:
                headers = ["IP", "Identity", "Model", "Version", "Status", "Last Seen"]
                ws.append(headers)

                # Style headers
                for col in range(1, len(headers) + 1):
                    ws.cell(row=1, column=col).font = Font(bold=True)
                    ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor="CCCCCC")

                for router in data["routers"]:
                    ws.append([
                        router.get("ip"),
                        router.get("identity"),
                        router.get("model"),
                        router.get("ros_version"),
                        "Online" if router.get("is_online") else "Offline",
                        router.get("last_seen")
                    ])

            wb.save(filepath)

            return {
                "id": filename,
                "report_type": request.report_type,
                "title": request.title or f"{request.report_type.title()} Report",
                "format": "excel",
                "status": "completed",
                "file_path": str(filepath),
                "file_size": filepath.stat().st_size,
                "download_url": f"/api/reports/download/{filename}",
                "created_at": datetime.utcnow().isoformat()
            }

        except ImportError:
            logger.warning("openpyxl not installed, Excel generation disabled")
            return self._generate_csv(request, data)

    def _generate_csv(self, request: ReportRequest, data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate CSV report"""
        filename = f"report_{request.report_type}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
        filepath = self.export_dir / filename

        with open(filepath, 'w', newline='') as f:
            writer = csv.writer(f)

            if "routers" in data:
                writer.writerow(["IP", "Identity", "Model", "Version", "Status", "Last Seen"])
                for router in data["routers"]:
                    writer.writerow([
                        router.get("ip"),
                        router.get("identity"),
                        router.get("model"),
                        router.get("ros_version"),
                        "Online" if router.get("is_online") else "Offline",
                        router.get("last_seen")
                    ])

        return {
            "id": filename,
            "report_type": request.report_type,
            "title": request.title or f"{request.report_type.title()} Report",
            "format": "csv",
            "status": "completed",
            "file_path": str(filepath),
            "file_size": filepath.stat().st_size,
            "download_url": f"/api/reports/download/{filename}",
            "created_at": datetime.utcnow().isoformat()
        }

    def _generate_json(self, request: ReportRequest, data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate JSON report"""
        filename = f"report_{request.report_type}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
        filepath = self.export_dir / filename

        with open(filepath, 'w') as f:
            json.dump(data, f, indent=2, default=str)

        return {
            "id": filename,
            "report_type": request.report_type,
            "title": request.title or f"{request.report_type.title()} Report",
            "format": "json",
            "status": "completed",
            "file_path": str(filepath),
            "file_size": filepath.stat().st_size,
            "download_url": f"/api/reports/download/{filename}",
            "created_at": datetime.utcnow().isoformat()
        }

    def get_report_file(self, filename: str) -> Optional[Path]:
        """Get report file path"""
        filepath = self.export_dir / filename
        if filepath.exists():
            return filepath
        return None

    def list_reports(self) -> List[Dict[str, Any]]:
        """List generated reports"""
        reports = []
        for filepath in self.export_dir.glob("report_*"):
            reports.append({
                "filename": filepath.name,
                "file_size": filepath.stat().st_size,
                "created_at": datetime.fromtimestamp(filepath.stat().st_mtime).isoformat(),
                "download_url": f"/api/reports/download/{filepath.name}"
            })
        return sorted(reports, key=lambda x: x["created_at"], reverse=True)

    def delete_report(self, filename: str) -> bool:
        """Delete a report file"""
        filepath = self.export_dir / filename
        if filepath.exists():
            filepath.unlink()
            return True
        return False
